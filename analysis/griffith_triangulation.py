#!/usr/bin/env python3
"""
Griffith et al. (2022) triangulation analysis for §5.8.
Derives per-participant metrics from raw 40-participant × 8-task data.

Data source: openICPSR project 183081
Primary: Excel file "Copy of FinalData.xlsx" (deposited raw data)
Cross-validation: Paper Table 4 (published time values)

Sites:
  MM = My Market (high-a11y grocery)      → maps to ~Base/High in our framework
  GG = Great Grocery (low-a11y grocery)    → maps to L3 (broken dropdown, aria-hidden)
  OU = Oakleaf University (high-a11y uni)  → maps to ~Base/High in our framework
  PU = Pinebranch University (low-a11y uni)→ maps to L3 (broken links, no landmarks)

Each participant did 2 tasks per site (task 1 and task 2).
Status: Y = completed, N = not completed
Time: in seconds

Usage:
  .venv/bin/python3 griffith_triangulation.py
"""

import numpy as np
import openpyxl
import datetime
import re
import os

TASK_ORDER = ['MM1', 'MM2', 'GG1', 'GG2', 'OU1', 'OU2', 'PU1', 'PU2']
EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'Copy of FinalData.xlsx')


def load_from_excel(path=EXCEL_PATH):
    """Load raw data from the deposited Excel file."""
    wb = openpyxl.load_workbook(path)
    ws = wb['Time and Task Performance']

    # Find data block start rows (rows where column A = 'MM1')
    block_starts = []
    for r in range(1, 100):
        val = ws.cell(row=r, column=1).value
        if val and str(val).strip() == 'MM1':
            block_starts.append(r)

    data = {}
    for block_row in block_starts:
        header_row = block_row - 2
        for col_start in [1, 5, 9, 13, 17]:
            pid_cell = ws.cell(row=header_row, column=col_start).value
            if pid_cell is None:
                continue
            matches = re.findall(r'P(\d+)', str(pid_cell).strip())
            if not matches:
                continue
            pid = int(matches[-1])
            data[pid] = {}
            for ti, task in enumerate(TASK_ORDER):
                row = block_row + ti
                status_val = ws.cell(row=row, column=col_start + 1).value
                time_val = ws.cell(row=row, column=col_start + 2).value

                success = (str(status_val).strip().upper() == 'Y') if status_val else False

                if isinstance(time_val, datetime.time):
                    # Excel stores M:SS as HH:MM:SS → hour=minutes, minute=seconds
                    time_secs = time_val.hour * 60 + time_val.minute
                elif isinstance(time_val, str):
                    m = re.search(r'(\d+):(\d+)', time_val)
                    time_secs = int(m.group(1)) * 60 + int(m.group(2)) if m else 0
                else:
                    time_secs = 0

                data[pid][task] = (success, time_secs)

    assert len(data) == 40, f"Expected 40 participants, got {len(data)}"
    return data


# Paper Table 4 times for cross-validation
PAPER_TABLE4_TIMES = {
    1: [82,40,48,246,83,45,578,520], 2: [125,59,100,164,77,85,202,150],
    3: [82,85,61,205,70,125,355,319], 4: [60,19,166,100,40,27,215,136],
    5: [93,40,135,341,61,136,307,189], 6: [77,34,58,98,30,22,85,100],
    7: [67,30,66,160,14,60,185,114], 8: [102,370,404,159,184,115,403,159],
    9: [54,51,60,79,38,81,47,33], 10: [105,36,61,95,38,52,153,300],
    11: [66,52,59,132,32,36,137,72], 12: [130,63,310,246,35,25,273,100],
    13: [76,40,37,140,100,80,235,120], 14: [32,30,32,151,35,28,195,131],
    15: [105,77,112,369,72,63,247,158], 16: [98,32,35,413,22,93,465,283],
    17: [58,63,56,196,70,32,328,226], 18: [99,135,125,99,62,75,186,126],
    19: [78,38,55,181,55,39,187,76], 20: [159,84,47,368,69,86,520,179],
    21: [63,47,514,418,153,54,338,97], 22: [88,265,474,313,56,65,385,392],
    23: [41,71,200,250,60,60,185,103], 24: [39,53,61,264,64,51,298,168],
    25: [50,95,49,167,36,45,233,210], 26: [22,55,44,188,55,23,144,181],
    27: [54,187,96,321,125,120,127,108], 28: [65,70,167,279,37,91,270,291],
    29: [65,33,120,82,56,173,165,235], 30: [60,56,207,84,51,53,224,184],
    31: [123,158,218,98,83,92,122,113], 32: [40,30,237,175,31,18,85,68],
    33: [59,64,158,376,45,60,205,120], 34: [64,71,182,345,358,197,506,430],
    35: [61,49,103,372,83,45,241,231], 36: [54,257,224,173,203,70,163,77],
    37: [121,187,288,362,81,91,428,230], 38: [47,64,33,172,44,21,117,149],
    39: [42,262,68,355,45,155,440,149], 40: [73,100,79,371,43,55,458,420],
}


def cross_validate(data):
    """Assert Excel data matches Paper Table 4 times (within 1s tolerance)."""
    mismatches = 0
    for pid in range(1, 41):
        for ti, task in enumerate(TASK_ORDER):
            if pid == 23 and task == 'GG1':
                continue  # ambiguous N/A entry
            excel_time = data[pid][task][1]
            paper_time = PAPER_TABLE4_TIMES[pid][ti]
            if abs(excel_time - paper_time) > 1:
                print(f"  MISMATCH P{pid} {task}: Excel={excel_time}s, Paper={paper_time}s")
                mismatches += 1
    assert mismatches == 0, f"{mismatches} time mismatches between Excel and Paper Table 4"
    print(f"  Cross-validation passed: 0 mismatches on 319 cells (P23-GG1 excluded as ambiguous)")


def run_analysis(raw):
    """Run all derived analyses and print results."""
    np.random.seed(42)

    # Per-site success rates
    print("\n--- Per-site success rates ---")
    sites = {'MM': ['MM1','MM2'], 'GG': ['GG1','GG2'], 'OU': ['OU1','OU2'], 'PU': ['PU1','PU2']}
    for site, tasks in sites.items():
        succ = sum(1 for pid in raw for t in tasks if raw[pid][t][0])
        total = sum(1 for pid in raw for t in tasks)
        print(f"  {site}: {succ}/{total} = {succ/total*100:.1f}%")

    # Per-participant time ratio (PU / OU)
    print("\n--- Time ratio (PU / OU) ---")
    ratios = []
    for pid in raw:
        ou = np.median([raw[pid]['OU1'][1], raw[pid]['OU2'][1]])
        pu = np.median([raw[pid]['PU1'][1], raw[pid]['PU2'][1]])
        ratios.append(pu / ou if ou > 0 else float('inf'))
    ratios = np.array(ratios)
    boot = [np.mean(np.random.choice(ratios, len(ratios), replace=True)) for _ in range(10000)]
    print(f"  Mean: {np.mean(ratios):.2f}, Median: {np.median(ratios):.2f}, SD: {np.std(ratios,ddof=1):.2f}")
    print(f"  95% bootstrap CI: [{np.percentile(boot,2.5):.2f}, {np.percentile(boot,97.5):.2f}]")

    # Per-participant success delta
    print("\n--- Success delta (high - low) ---")
    deltas = []
    for pid in raw:
        high = sum(1 for t in ['MM1','MM2','OU1','OU2'] if raw[pid][t][0]) / 4.0
        low = sum(1 for t in ['GG1','GG2','PU1','PU2'] if raw[pid][t][0]) / 4.0
        deltas.append(high - low)
    deltas = np.array(deltas)
    boot_d = [np.mean(np.random.choice(deltas, len(deltas), replace=True)) for _ in range(10000)]
    print(f"  Mean: {np.mean(deltas):.3f}, SD: {np.std(deltas,ddof=1):.3f}")
    print(f"  95% bootstrap CI: [{np.percentile(boot_d,2.5):.3f}, {np.percentile(boot_d,97.5):.3f}]")
    print(f"  All >= 0: {all(d >= 0 for d in deltas)}")

    # Variance decomposition
    print("\n--- Variance decomposition ---")
    site_names = ['MM', 'GG', 'OU', 'PU']
    obs = []
    for pid in raw:
        for si, site in enumerate(site_names):
            for task in sites[site]:
                obs.append((si, pid, 1.0 if raw[pid][task][0] else 0.0))
    obs = np.array(obs)
    gm = np.mean(obs[:,2])
    sm = {s: np.mean(obs[obs[:,0]==si, 2]) for si, s in enumerate(site_names)}
    pm = {pid: np.mean(obs[obs[:,1]==pid, 2]) for pid in raw}
    ss_t = np.sum((obs[:,2] - gm)**2)
    ss_s = sum(np.sum(obs[:,0]==si) * (sm[s]-gm)**2 for si,s in enumerate(site_names))
    ss_p = sum(np.sum(obs[:,1]==pid) * (pm[pid]-gm)**2 for pid in raw)
    ss_r = ss_t - ss_s - ss_p
    print(f"  Grand mean: {gm:.3f}")
    print(f"  Site means: {', '.join(f'{s}={sm[s]:.3f}' for s in site_names)}")
    print(f"  SS_site: {ss_s:.2f} ({ss_s/ss_t*100:.1f}%)")
    print(f"  SS_participant: {ss_p:.2f} ({ss_p/ss_t*100:.1f}%)")
    print(f"  SS_residual: {ss_r:.2f} ({ss_r/ss_t*100:.1f}%)")

    # Split reporting
    print("\n--- Split reporting (high / GG / PU) ---")
    high_s = [raw[pid][t][0] for pid in raw for t in ['MM1','MM2','OU1','OU2']]
    print(f"  High-a11y (MM+OU): {sum(high_s)}/{len(high_s)} = {sum(high_s)/len(high_s)*100:.1f}%")
    gg_s = [raw[pid][t][0] for pid in raw for t in ['GG1','GG2']]
    print(f"  GG (L3-light): {sum(gg_s)}/{len(gg_s)} = {sum(gg_s)/len(gg_s)*100:.1f}%")
    pu_s = [raw[pid][t][0] for pid in raw for t in ['PU1','PU2']]
    print(f"  PU (L3-heavy): {sum(pu_s)}/{len(pu_s)} = {sum(pu_s)/len(pu_s)*100:.1f}%")

    # Time inflation
    high_t = [raw[pid][t][1] for pid in raw for t in ['MM1','MM2','OU1','OU2']]
    low_t = [raw[pid][t][1] for pid in raw for t in ['GG1','GG2','PU1','PU2']]
    print(f"\n  High-a11y median time: {np.median(high_t):.0f}s")
    print(f"  L3-low median time: {np.median(low_t):.0f}s")
    print(f"  Time inflation: {np.median(low_t)/np.median(high_t):.1f}x")


if __name__ == '__main__':
    print("Loading from Excel...")
    data = load_from_excel()
    print(f"Loaded {len(data)} participants\n")

    print("Cross-validating against Paper Table 4...")
    cross_validate(data)

    run_analysis(data)
